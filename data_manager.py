from openpyxl import load_workbook
from datetime import datetime, timedelta
from typing import Dict, List
import os

class DataManager:
    """Handles all Excel data loading using only openpyxl"""
    
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self._cache = {}
        self._last_load = None
        self.cache_duration = timedelta(minutes=5)
    
    def _should_reload(self) -> bool:
        """Check if cache should be refreshed"""
        if not self._last_load:
            return True
        return datetime.now() - self._last_load > self.cache_duration
    
    def _sheet_to_dict_list(self, worksheet) -> List[Dict]:
        """Convert worksheet to list of dictionaries"""
        data = []
        headers = []
        
        for i, row in enumerate(worksheet.iter_rows(values_only=True)):
            if i == 0:
                headers = list(row)
            else:
                row_dict = {}
                for j, value in enumerate(row):
                    if j < len(headers):
                        row_dict[headers[j]] = value
                data.append(row_dict)
        
        return data
    
    def load_data(self, force_reload: bool = False) -> Dict[str, List[Dict]]:
        """Load all Excel sheets into dictionaries with caching"""
        if not force_reload and not self._should_reload() and self._cache:
            return self._cache
        
        try:
            workbook = load_workbook(self.excel_path, data_only=True)
            
            # Load all sheets
            employees = self._sheet_to_dict_list(workbook['Employees'])
            projects = self._sheet_to_dict_list(workbook['Projects'])
            assignments = self._sheet_to_dict_list(workbook['Project_Assignments'])
            
            # Try to load Daily_Billing, generate if missing
            try:
                daily_billing = self._sheet_to_dict_list(workbook['Daily_Billing'])
            except:
                daily_billing = self._generate_daily_billing(assignments)
            
            # Convert date strings to datetime objects
            for emp in employees:
                if emp.get('Joining_Date'):
                    emp['Joining_Date'] = self._parse_date(emp['Joining_Date'])
            
            for proj in projects:
                if proj.get('Start_Date'):
                    proj['Start_Date'] = self._parse_date(proj['Start_Date'])
                if proj.get('End_Date'):
                    proj['End_Date'] = self._parse_date(proj['End_Date'])
            
            for assign in assignments:
                if assign.get('Billing_Start_Date'):
                    assign['Billing_Start_Date'] = self._parse_date(assign['Billing_Start_Date'])
                if assign.get('Billing_End_Date'):
                    assign['Billing_End_Date'] = self._parse_date(assign['Billing_End_Date'])
            
            for billing in daily_billing:
                if billing.get('Date'):
                    billing['Date'] = self._parse_date(billing['Date'])
            
            self._cache = {
                'employees': employees,
                'projects': projects,
                'assignments': assignments,
                'daily_billing': daily_billing
            }
            
            self._last_load = datetime.now()
            workbook.close()
            return self._cache
            
        except Exception as e:
            raise Exception(f"Error loading Excel file: {str(e)}")
    
    def _parse_date(self, date_value):
        """Parse date from various formats"""
        if isinstance(date_value, datetime):
            return date_value
        if isinstance(date_value, str):
            try:
                return datetime.strptime(date_value, '%Y-%m-%d')
            except:
                try:
                    return datetime.strptime(date_value, '%Y-%m-%d %H:%M:%S')
                except:
                    return datetime.now()
        return date_value
    
    def _generate_daily_billing(self, assignments: List[Dict]) -> List[Dict]:
        """Generate daily billing records from assignments"""
        records = []
        
        for assignment in assignments:
            start = self._parse_date(assignment.get('Billing_Start_Date'))
            end = self._parse_date(assignment.get('Billing_End_Date'))
            employee_id = assignment.get('Employee_ID')
            
            if not start or not end or not employee_id:
                continue
            
            current_date = start
            while current_date <= end:
                # Only weekdays
                if current_date.weekday() < 5:
                    records.append({
                        'Employee_ID': employee_id,
                        'Date': current_date,
                        'Is_Billed': 'Yes'
                    })
                current_date += timedelta(days=1)
        
        return records
    
    def get_employees(self, tool: str = None) -> List[Dict]:
        """Get employees, optionally filtered by tool"""
        data = self.load_data()
        employees = data['employees'].copy()
        
        if tool:
            employees = [e for e in employees if e.get('Tool') == tool]
        
        return employees
    
    def get_projects(self, status: str = None, tool: str = None) -> List[Dict]:
        """Get projects filtered by status and/or tool"""
        data = self.load_data()
        projects = data['projects'].copy()
        
        if status:
            projects = [p for p in projects if p.get('Project_Status') == status]
        if tool:
            projects = [p for p in projects if p.get('Tool') == tool]
        
        return projects
    
    def get_project_members(self, project_id: int) -> List[Dict]:
        """Get all members assigned to a project with their details"""
        data = self.load_data()
        
        # Get assignments for this project
        project_assignments = [a for a in data['assignments'] 
                              if a.get('Project_ID') == project_id]
        
        # Merge with employee data
        members = []
        for assignment in project_assignments:
            emp_id = assignment.get('Employee_ID')
            employee = next((e for e in data['employees'] 
                           if e.get('Employee_ID') == emp_id), None)
            
            if employee:
                member = {**assignment, **employee}
                members.append(member)
        
        return members
    
    def get_employee_projects(self, employee_id: int, year: int = None) -> List[Dict]:
        """Get all projects for an employee"""
        data = self.load_data()
        
        # Get employee assignments
        employee_assignments = [a for a in data['assignments'] 
                               if a.get('Employee_ID') == employee_id]
        
        # Merge with project data
        projects = []
        for assignment in employee_assignments:
            proj_id = assignment.get('Project_ID')
            project = next((p for p in data['projects'] 
                          if p.get('Project_ID') == proj_id), None)
            
            if project:
                merged = {**assignment, **project}
                
                if year:
                    start_year = merged.get('Start_Date').year if merged.get('Start_Date') else None
                    end_year = merged.get('End_Date').year if merged.get('End_Date') else None
                    
                    if start_year == year or end_year == year:
                        projects.append(merged)
                else:
                    projects.append(merged)
        
        return projects
    
    def get_billing_records(self, employee_id: int = None, 
                           start_date: datetime = None, 
                           end_date: datetime = None) -> List[Dict]:
        """Get billing records with optional filters"""
        data = self.load_data()
        billing = data['daily_billing'].copy()
        
        if employee_id:
            billing = [b for b in billing if b.get('Employee_ID') == employee_id]
        
        if start_date:
            billing = [b for b in billing if b.get('Date') and b['Date'] >= start_date]
        
        if end_date:
            billing = [b for b in billing if b.get('Date') and b['Date'] <= end_date]
        
        return billing
