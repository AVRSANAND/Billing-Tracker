# ============================================================================
# FILE: create_sample_data.py (NO PANDAS VERSION)
# ============================================================================
from openpyxl import Workbook
from datetime import datetime, timedelta
import os
import random

def create_sample_excel():
    """Generate sample Excel file with realistic data using openpyxl only"""
    
    # Create data directory
    os.makedirs('data', exist_ok=True)
    
    # Create workbook
    wb = Workbook()
    
    # Sheet 1: Employees
    ws_employees = wb.active
    ws_employees.title = 'Employees'
    ws_employees.append(['Employee_ID', 'Employee_Name', 'Tool', 'Role', 'Joining_Date'])
    
    employee_names = [
        'Alice Johnson', 'Bob Smith', 'Charlie Brown', 'Diana Prince', 'Eve Adams',
        'Frank Miller', 'Grace Lee', 'Henry Ford', 'Iris Wang', 'Jack Ryan',
        'Kate Wilson', 'Leo Martinez', 'Maria Garcia', 'Nathan Drake', 'Olivia Moore',
        'Peter Parker', 'Quinn Taylor', 'Rachel Green', 'Sam Fisher', 'Tina Turner',
        'Uma Thurman', 'Victor Hugo', 'Wendy Williams', 'Xavier Lopez', 'Yara Shah',
        'Zack Morris', 'Amy Chen', 'Brian Cox', 'Carla Diaz', 'David Kim'
    ]
    
    roles = ['Senior Engineer', 'Engineer', 'Lead', 'Engineer', 'Senior Engineer',
             'Architect', 'Engineer', 'Lead', 'Senior Engineer', 'Engineer']
    
    for i in range(30):
        emp_id = i + 1
        name = employee_names[i]
        tool = ['CAP360', 'BREAD', 'DCC'][i // 10]
        role = roles[i % 10]
        joining_date = datetime(2023, random.randint(1, 12), random.randint(1, 28))
        
        ws_employees.append([emp_id, name, tool, role, joining_date.strftime('%Y-%m-%d')])
    
    # Sheet 2: Projects
    ws_projects = wb.create_sheet('Projects')
    ws_projects.append(['Project_ID', 'Project_Name', 'Tool', 'Project_Status', 'Start_Date', 'End_Date'])
    
    projects_data = [
        (1, 'Digital Transformation Portal', 'CAP360', 'Ongoing', datetime(2024, 1, 15), datetime(2024, 12, 31)),
        (2, 'Cloud Migration Phase 2', 'CAP360', 'Ongoing', datetime(2024, 3, 1), datetime(2024, 12, 31)),
        (3, 'Analytics Dashboard', 'CAP360', 'Completed', datetime(2023, 10, 1), datetime(2024, 3, 31)),
        (4, 'Mobile App Redesign', 'BREAD', 'Ongoing', datetime(2024, 2, 1), datetime(2024, 12, 31)),
        (5, 'API Gateway Implementation', 'BREAD', 'Upcoming', datetime(2025, 1, 1), datetime(2025, 6, 30)),
        (6, 'Security Enhancement', 'BREAD', 'Ongoing', datetime(2024, 4, 1), datetime(2024, 12, 31)),
        (7, 'Data Warehouse Upgrade', 'BREAD', 'Completed', datetime(2023, 8, 1), datetime(2024, 2, 28)),
        (8, 'Customer Portal V2', 'DCC', 'Ongoing', datetime(2024, 5, 1), datetime(2024, 12, 31)),
        (9, 'IoT Integration Platform', 'DCC', 'Upcoming', datetime(2025, 2, 1), datetime(2025, 8, 31)),
        (10, 'AI/ML Model Deployment', 'DCC', 'Ongoing', datetime(2024, 6, 1), datetime(2024, 12, 31)),
        (11, 'Legacy System Modernization', 'CAP360', 'Completed', datetime(2023, 6, 1), datetime(2024, 1, 31)),
        (12, 'DevOps Pipeline', 'BREAD', 'Ongoing', datetime(2024, 7, 1), datetime(2024, 12, 31)),
        (13, 'E-commerce Platform', 'DCC', 'Ongoing', datetime(2024, 8, 1), datetime(2024, 12, 31)),
        (14, 'CRM Integration', 'CAP360', 'Upcoming', datetime(2025, 1, 15), datetime(2025, 7, 31)),
        (15, 'Blockchain POC', 'BREAD', 'Completed', datetime(2023, 12, 1), datetime(2024, 5, 31)),
        (16, 'Microservices Architecture', 'DCC', 'Ongoing', datetime(2024, 9, 1), datetime(2024, 12, 31)),
        (17, 'Testing Automation', 'CAP360', 'Completed', datetime(2023, 11, 1), datetime(2024, 4, 30)),
        (18, 'Performance Optimization', 'BREAD', 'Ongoing', datetime(2024, 10, 1), datetime(2024, 12, 31)),
        (19, 'Compliance Management', 'DCC', 'Upcoming', datetime(2025, 3, 1), datetime(2025, 9, 30)),
        (20, 'Digital Marketing Tools', 'CAP360', 'Completed', datetime(2023, 9, 1), datetime(2024, 3, 31)),
    ]
    
    for proj in projects_data:
        proj_id, name, tool, status, start, end = proj
        ws_projects.append([proj_id, name, tool, status, start.strftime('%Y-%m-%d'), end.strftime('%Y-%m-%d')])
    
    # Sheet 3: Project_Assignments
    ws_assignments = wb.create_sheet('Project_Assignments')
    ws_assignments.append(['Employee_ID', 'Project_ID', 'Billing_Start_Date', 'Billing_End_Date', 'Billability_Percentage'])
    
    assignments = []
    for proj in projects_data:
        proj_id, _, proj_tool, _, proj_start, proj_end = proj
        
        # Get employees from same tool (1-10 = CAP360, 11-20 = BREAD, 21-30 = DCC)
        if proj_tool == 'CAP360':
            emp_pool = list(range(1, 11))
        elif proj_tool == 'BREAD':
            emp_pool = list(range(11, 21))
        else:
            emp_pool = list(range(21, 31))
        
        # Assign 3-5 random employees
        num_assigned = random.randint(3, 5)
        assigned_employees = random.sample(emp_pool, min(num_assigned, len(emp_pool)))
        
        for emp_id in assigned_employees:
            start_offset = random.randint(0, 30)
            assignment_start = proj_start + timedelta(days=start_offset)
            assignment_end = min(
                assignment_start + timedelta(days=random.randint(60, 180)),
                proj_end
            )
            billability = random.choice([80, 85, 90, 95, 100])
            
            assignments.append((emp_id, proj_id, assignment_start, assignment_end, billability))
            ws_assignments.append([
                emp_id, proj_id, 
                assignment_start.strftime('%Y-%m-%d'), 
                assignment_end.strftime('%Y-%m-%d'), 
                billability
            ])
    
    # Sheet 4: Daily_Billing
    ws_billing = wb.create_sheet('Daily_Billing')
    ws_billing.append(['Employee_ID', 'Date', 'Is_Billed'])
    
    billing_count = 0
    for emp_id, proj_id, start, end, _ in assignments:
        current_date = start
        end_date = min(end, datetime.now())
        
        while current_date <= end_date:
            # Only weekdays, random 85% billing rate
            if current_date.weekday() < 5 and random.random() < 0.85:
                ws_billing.append([emp_id, current_date.strftime('%Y-%m-%d'), 'Yes'])
                billing_count += 1
            current_date += timedelta(days=1)
    
    # Save workbook
    excel_path = 'data/billability_data.xlsx'
    wb.save(excel_path)
    
    print(f"✅ Sample Excel file created: {excel_path}")
    print(f"   - 30 Employees")
    print(f"   - 20 Projects")
    print(f"   - {len(assignments)} Project Assignments")
    print(f"   - {billing_count} Daily Billing Records")

if __name__ == '__main__':
    create_sample_excel()